from __future__ import annotations

import io
from datetime import date
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.errors import raise_app_error
from app.models import Member, Reservation, ReservationCycle, ReservationStatus, Slot
from app.services.cycle import get_admin_view_cycle, get_slot_times, week_dates

TEMPLATE_PATH = Path(__file__).resolve().parents[2] / "data" / "applicant_export_template.xlsx"

DATE_HEADER_ROW = 5
TIME_COL = 2  # B
FIRST_DATE_COL = 3  # C
FIRST_TIME_ROW = 6
DATA_FIRST_ROW = 4
DATA_LAST_ROW = 9
DATA_FIRST_COL = 2  # B
DATA_LAST_COL = 7  # G

CENTER_ALIGN = Alignment(horizontal="center", vertical="center")
TOP_ALIGN = Alignment(horizontal="center", vertical="top")


def _date_header_label(d: date) -> str:
    return f"{d.month}월{d.day}일"


def _week_filename_label(week_start: date, week_end: date) -> str:
    return f"{_date_header_label(week_start)}-{_date_header_label(week_end)}"


def _time_label(start: str, end: str) -> str:
    return f"{start}~{end}"


async def _build_time_row_map(db: AsyncSession) -> dict[str, int]:
    slot_times = await get_slot_times(db)
    return {
        _time_label(slot["s"], slot["e"]): FIRST_TIME_ROW + idx
        for idx, slot in enumerate(slot_times)
    }


def _apply_cell_alignment(ws, dates_count: int, time_rows: set[int]) -> None:
    for row in range(DATA_FIRST_ROW, DATA_LAST_ROW + 1):
        for col in range(DATA_FIRST_COL, DATA_LAST_COL + 1):
            ws.cell(row=row, column=col).alignment = CENTER_ALIGN

    for row in time_rows:
        for day_idx in range(dates_count):
            ws.cell(row=row, column=FIRST_DATE_COL + day_idx).alignment = TOP_ALIGN


async def export_confirmed_applicants_xlsx(
    db: AsyncSession,
    cycle_id: int | None = None,
) -> tuple[bytes, str]:
    if not cycle_id:
        cycle = await get_admin_view_cycle(db)
        cycle_id = cycle.id if cycle else None
    if not cycle_id:
        raise_app_error("NOT_FOUND", 404)

    cycle = await db.get(ReservationCycle, cycle_id)
    if not cycle:
        raise_app_error("NOT_FOUND", 404)

    if not TEMPLATE_PATH.is_file():
        raise_app_error("NOT_FOUND", 500)

    week_start = cycle.target_week_start
    week_end = cycle.target_week_end
    dates = [date.fromisoformat(d) for d in week_dates(week_start)]
    row_map = await _build_time_row_map(db)

    result = await db.execute(
        select(Slot, Reservation, Member)
        .join(Reservation, Reservation.slot_id == Slot.id)
        .join(Member, Member.id == Reservation.member_id)
        .where(
            Slot.cycle_id == cycle_id,
            Reservation.status == ReservationStatus.CONFIRMED,
        )
        .order_by(Slot.slot_date, Slot.time_index)
    )

    confirmed_by_cell: dict[tuple[int, int], str] = {}
    for slot, _reservation, member in result.all():
        try:
            day_idx = dates.index(slot.slot_date)
        except ValueError:
            continue
        col = FIRST_DATE_COL + day_idx
        time_label = _time_label(
            slot.start_time.strftime("%H:%M"),
            slot.end_time.strftime("%H:%M"),
        )
        row = row_map.get(time_label)
        if row is None:
            continue
        confirmed_by_cell[(row, col)] = member.name

    wb = load_workbook(TEMPLATE_PATH)
    ws = wb.active
    ws.title = "확정자"

    for day_idx, d in enumerate(dates):
        ws.cell(row=DATE_HEADER_ROW, column=FIRST_DATE_COL + day_idx, value=_date_header_label(d))

    for time_label, row in row_map.items():
        ws.cell(row=row, column=TIME_COL, value=time_label)
        for day_idx in range(len(dates)):
            col = FIRST_DATE_COL + day_idx
            name = confirmed_by_cell.get((row, col))
            if name:
                ws.cell(row=row, column=col, value=name)

    _apply_cell_alignment(ws, len(dates), set(row_map.values()))

    buf = io.BytesIO()
    wb.save(buf)
    filename = f"헬스키퍼_{_week_filename_label(week_start, week_end)}_신청자 목록.xlsx"
    return buf.getvalue(), filename
